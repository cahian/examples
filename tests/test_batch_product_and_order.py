from io import BytesIO

import pytest
from django.core.exceptions import ObjectDoesNotExist
from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from core.api.views.batch_product_and_order import (
	EXCEL_MAX_ROWS,
	ProductAndSaleOrderBase,
	ProductAndSaleOrderTemplateGenerator,
	SpreadsheetCoordinatesManager,
)
from core.models import Website


class TestSpreadsheetCoordinatesManager:
	def test_initialization_empty(self):
		"""**Test initialization with empty data**"""
		manager = SpreadsheetCoordinatesManager({})
		assert manager.data == {}

	def test_initialization_with_data(self):
		"""**Test initialization with predefined data**"""
		initial_data = {'column1': 1, 'column2': 2}
		manager = SpreadsheetCoordinatesManager(initial_data)
		assert manager.data == initial_data

	def test_get_column_existing(self):
		"""**Test getting an existing column**"""
		manager = SpreadsheetCoordinatesManager({'column1': 1})
		assert manager.get_column_index('column1') == 1

	def test_get_column_non_existing(self):
		"""**Test getting a non-existing column raises KeyError**"""
		manager = SpreadsheetCoordinatesManager({'column1': 1})
		with pytest.raises(KeyError):
			manager.get_column_index('non_existing_column')

	def test_add_column(self):
		"""**Test adding a new column**"""
		manager = SpreadsheetCoordinatesManager({})
		manager.add_column_index('column1', 1)
		assert manager.data['column1'] == 1

	def test_add_existing_column_raises_error(self):
		"""**Test adding an existing column raises ValueError**"""
		manager = SpreadsheetCoordinatesManager({'column1': 1})
		with pytest.raises(ValueError, match='Column column1 already exists'):
			manager.add_column_index('column1', 2)

	def test_update_existing_column(self):
		"""**Test updating an existing column**"""
		manager = SpreadsheetCoordinatesManager({'column1': 2})
		manager.update_column_offset('column1', 3)
		assert manager.data['column1'] == 5

	def test_update_column_to_none(self):
		"""**Test updating a column to None**"""
		manager = SpreadsheetCoordinatesManager({'column1': 5})
		manager.update_column_offset('column1', None)
		assert manager.data['column1'] is None

	def test_update_non_existing_column_raises_error(self):
		"""**Test updating a non-existing column raises KeyError**"""
		manager = SpreadsheetCoordinatesManager({})
		with pytest.raises(KeyError):
			manager.update_column_offset('non_existing_column', 5)

	def test_get_coordinate_valid(self):
		"""**Test getting coordinate for an existing column and row**"""
		manager = SpreadsheetCoordinatesManager(
			{'column1': 3}
		)  # C is the third column
		row = 2
		expected_coordinate = f'{get_column_letter(3)}{row}'  # Should be C2
		assert manager.get_coordinate('column1', row) == expected_coordinate

	def test_get_coordinate_non_existing(self):
		"""**Test getting coordinate for a non-existing column returns None**"""
		manager = SpreadsheetCoordinatesManager({'column1': 3})
		with pytest.raises(KeyError):
			manager.get_coordinate('non_existing_column', 2)

	def test_get_coordinate_range_valid(self):
		"""**Test getting coordinate range for an existing column**"""
		manager = SpreadsheetCoordinatesManager(
			{'column1': 4}
		)  # D is the fourth column
		expected_range = f'{get_column_letter(4)}3:{get_column_letter(4)}{EXCEL_MAX_ROWS}'  # D3:D{EXCEL_MAX_ROWS}
		assert manager.get_coordinate_range('column1') == expected_range

	def test_get_coordinate_range_non_existing(self):
		"""**Test getting coordinate range for a non-existing column returns None**"""
		manager = SpreadsheetCoordinatesManager({})
		with pytest.raises(KeyError):
			manager.get_coordinate_range('non_existing_column')

	def test_add_multiple_columns_and_get_coordinates(self):
		"""**Test adding multiple columns and retrieving their coordinates**"""
		manager = SpreadsheetCoordinatesManager({})

		# Add multiple columns
		columns_to_add = {
			'A': 1,
			'B': 2,
			'C': 3,
			'D': 4,
			'E': 5,
			'F': None,  # Column without index
			'G': None,  # Another column without index
			'H': -2,  # Invalid index (negative)
		}

		for key, value in columns_to_add.items():
			if value is None:
				with pytest.raises(KeyError):
					manager.get_coordinate(key, 10)
			else:
				manager.add_column_index(key, value)
				if value <= 0:
					with pytest.raises(ValueError):
						manager.get_coordinate(key, 10)
				else:
					assert (
						manager.get_coordinate(key, 10)
						== f'{get_column_letter(value)}10'
					)

	def test_update_multiple_columns(self):
		"""**Test updating multiple columns in various scenarios**"""
		initial_data = {'A': 5, 'B': None}

		# Initialize with some data
		manager = SpreadsheetCoordinatesManager(initial_data)

		# Update existing column
		manager.update_column_offset('A', 5)  # Should become A:10

		# Check if column A is updated
		assert initial_data['A'] == 10

		with pytest.raises(KeyError):
			# Trying to update non-existing column
			manager.update_column_offset('C', 5)

	def test_delete_and_read_columns(self):
		"""**Test deleting and re-adding columns**"""
		initial_data = {'A': 5}
		manager = SpreadsheetCoordinatesManager(initial_data)

		# Delete the existing column A
		del manager.data['A']

		with pytest.raises(KeyError):
			# Check if key A raises KeyError after deletion
			_ = manager.get_column_index('A')

		# Re-add the same column A
		manager.add_column_index('A', 7)
		assert manager.get_column_index('A') == 7


@pytest.mark.django_db
class TestProductAndSaleOrderBase:
	@pytest.mark.parametrize(
		'indirect_company, expected_service_segment, expected_catalog, expected_product_code, expected_order_code',
		[
			(
				'company_no_segment_no_catalog_no_codes',
				False,
				False,
				False,
				False,
			),
			(
				'company_no_segment_no_catalog_order_code',
				False,
				False,
				False,
				True,
			),
			(
				'company_no_segment_no_catalog_product_code',
				False,
				False,
				True,
				False,
			),
			(
				'company_no_segment_no_catalog_both_codes',
				False,
				False,
				True,
				True,
			),
			(
				'company_no_segment_with_catalog_no_codes',
				False,
				True,
				False,
				False,
			),
			(
				'company_no_segment_with_catalog_order_code',
				False,
				True,
				False,
				True,
			),
			(
				'company_no_segment_with_catalog_product_code',
				False,
				True,
				True,
				False,
			),
			(
				'company_no_segment_with_catalog_both_codes',
				False,
				True,
				True,
				True,
			),
			(
				'company_with_service_segment_no_catalog_no_codes',
				True,
				False,
				False,
				False,
			),
			(
				'company_with_service_segment_no_catalog_order_code',
				True,
				False,
				False,
				True,
			),
			(
				'company_with_service_segment_no_catalog_product_code',
				True,
				False,
				True,
				False,
			),
			(
				'company_with_service_segment_no_catalog_both_codes',
				True,
				False,
				True,
				True,
			),
			(
				'company_with_service_segment_with_catalog_no_codes',
				True,
				True,
				False,
				False,
			),
			(
				'company_with_service_segment_with_catalog_order_code',
				True,
				True,
				False,
				True,
			),
			(
				'company_with_service_segment_with_catalog_product_code',
				True,
				True,
				True,
				False,
			),
			(
				'company_with_service_segment_with_all_codes',
				True,
				True,
				True,
				True,
			),
		],
		indirect=['indirect_company'],
	)
	def test_initialization_success(
		self,
		indirect_company,
		expected_catalog,
		expected_product_code,
		expected_order_code,
		expected_service_segment,
	):
		"""**Test successful initialization with various company configurations**"""
		company = indirect_company
		base = ProductAndSaleOrderBase(company.company_name)

		assert base.has_catalog == expected_catalog
		assert base.has_product_code == expected_product_code
		assert base.has_order_code == expected_order_code
		assert base.is_service_segment == expected_service_segment

	def test_initialization_invalid_company(self):
		"""**Test initialization raises ValueError with an invalid company name**"""
		with pytest.raises(ObjectDoesNotExist):
			ProductAndSaleOrderBase('invalid_company')

	def test_initialization_with_none_values(self):
		"""**Test behavior when None values are passed during initialization**"""
		with pytest.raises(ObjectDoesNotExist):
			ProductAndSaleOrderBase(None)

	def test_product_and_sale_order_base_methods_exist(self, company):
		"""**Verify that all expected methods exist in the class**"""
		base = ProductAndSaleOrderBase(company.company_name)
		expected_methods = ['get_template']
		for method in expected_methods:
			assert hasattr(
				base, method
			), f'Method {method} does not exist in ProductAndSaleOrderBase'


@pytest.mark.django_db
class TestProductAndSaleOrderTemplateGenerator:
	@pytest.mark.parametrize(
		'indirect_company',
		[
			'company_no_segment_no_catalog_no_codes',
			'company_no_segment_no_catalog_order_code',
			'company_no_segment_no_catalog_product_code',
			'company_no_segment_no_catalog_both_codes',
			'company_with_service_segment_no_catalog_product_code',
			'company_with_service_segment_no_catalog_both_codes',
		],
		indirect=True,
	)
	def test_get_template_raises_value_error(self, indirect_company, user):
		"""**Test that ValueError is raised for specific company configurations**"""
		company = indirect_company
		generator = ProductAndSaleOrderTemplateGenerator(
			user, company.company_name
		)

		with pytest.raises(ValueError):
			generator.get_template()

	@pytest.mark.parametrize(
		'indirect_company',
		[
			'company_no_segment_with_catalog_no_codes',
			'company_no_segment_with_catalog_order_code',
			'company_no_segment_with_catalog_product_code',
			'company_no_segment_with_catalog_both_codes',
			'company_with_service_segment_no_catalog_no_codes',
			'company_with_service_segment_no_catalog_order_code',
			'company_with_service_segment_with_catalog_no_codes',
			'company_with_service_segment_with_catalog_order_code',
			'company_with_service_segment_with_catalog_product_code',
			'company_with_service_segment_with_all_codes',
		],
		indirect=True,
	)
	def test_get_template_success(self, indirect_company, user):
		"""**Test if templates are generated successfully for all configurations**"""
		company = indirect_company
		generator = ProductAndSaleOrderTemplateGenerator(
			user, company.company_name
		)

		response = generator.get_template()

		assert isinstance(response, HttpResponse)
		assert response.status_code == 200
		assert response['Content-Type'] == 'application/ms-excel'

		# Load the workbook
		wb = load_workbook(BytesIO(response.content))

		# Validate expected sheets
		active_sheet = wb.active
		assert active_sheet.title in ['Produtos', 'Pedidos']

	@pytest.mark.parametrize(
		'indirect_company',
		[
			'company_with_service_segment_no_catalog_no_codes',
			'company_with_service_segment_no_catalog_order_code',
		],
		indirect=True,
	)
	def test_missing_catalog_removes_product_sheet(
		self,
		indirect_company,
		user,
	):
		"""**Test that the product sheet is removed if the company lacks a catalog**"""
		company = indirect_company
		generator = ProductAndSaleOrderTemplateGenerator(
			user, company.company_name
		)

		response = generator.get_template()
		wb = load_workbook(BytesIO(response.content))

		assert 'Produtos' not in wb.sheetnames

	def test_invalid_company_raises_error(self, user):
		"""**Test that an invalid company raises an exception**"""
		with pytest.raises(ObjectDoesNotExist):
			ProductAndSaleOrderTemplateGenerator(
				user, 'invalid_company_name'
			).get_template()

	@pytest.mark.parametrize(
		'indirect_company',
		[
			'company_no_segment_with_catalog_no_codes',
			'company_no_segment_with_catalog_order_code',
			'company_no_segment_with_catalog_product_code',
			'company_no_segment_with_catalog_both_codes',
			'company_with_service_segment_no_catalog_no_codes',
			'company_with_service_segment_no_catalog_order_code',
			'company_with_service_segment_with_catalog_no_codes',
			'company_with_service_segment_with_catalog_order_code',
			'company_with_service_segment_with_catalog_product_code',
			'company_with_service_segment_with_all_codes',
		],
		indirect=True,
	)
	def test_sku_headers_specific_behavior(
		self,
		indirect_company,
		user,
		service_segment,
	):
		"""**Test product headers in templates based on company configs**"""
		company = indirect_company
		generator = ProductAndSaleOrderTemplateGenerator(
			user, company.company_name
		)

		response = generator.get_template()
		wb = load_workbook(BytesIO(response.content))

		# Check if certain columns are renamed or removed
		if 'Produtos' in wb.sheetnames:
			product_sheet = wb['Produtos']
			headers = [cell.value for cell in product_sheet[2]]
			if not company.has_catalog:
				assert False, 'Produtos sheet should not exist for companies without a catalog'
			else:
				if company.spreadsheet_integration_configuration.has_product_code:
					if company.segment.name == service_segment.name:
						assert (
							'Código do Produto *' in headers
						), f'Código do Produto * should exist if product code is enabled: {headers}'
						assert (
							'SKU *' not in headers
						), f'SKU * should exist if product code is enabled: {headers}'
					else:
						assert (
							'Código do Produto *' not in headers
						), f'Código do Produto * should exist if product code is enabled: {headers}'
						assert (
							'SKU *' in headers
						), f'SKU * should exist if product code is enabled: {headers}'
					assert (
						'Nome do Produto *' not in headers
					), f'Nome do Produto * should not exist if SKU * is enabled: {headers}'
				else:
					assert (
						'Código do Produto *' not in headers
					), f'Código do Produto * should exist if product code is enabled: {headers}'
					assert (
						'SKU *' not in headers
					), f'SKU * should exist if product code is enabled: {headers}'
					assert (
						'Nome do Produto *' in headers
					), f'Nome do Produto * should exist if product code is disabled: {headers}'

		if 'Pedidos' in wb.sheetnames:
			order_sheet = wb['Pedidos']
			headers = [cell.value for cell in order_sheet[2]]

			if company.has_catalog:
				if company.spreadsheet_integration_configuration.has_product_code:
					if company.segment.name == service_segment.name:
						assert (
							'Código do Produto *' in headers
						), f'Código do Produto * should exist for catalog with product code: {headers}'
						assert (
							'SKU Variante *' not in headers
						), f'SKU Variante * should not exist without a catalog: {headers}'
					else:
						assert (
							'Código do Produto *' not in headers
						), f'Código do Produto * should exist for catalog with product code: {headers}'
						assert (
							'SKU Variante *' in headers
						), f'SKU Variante * should not exist without a catalog: {headers}'
					assert (
						'Nome do Produto *' not in headers
					), f'Nome do Produto * should replace SKU Variante * for catalog without product code: {headers}'
				else:
					assert (
						'Código do Produto *' not in headers
					), f'Código do Produto * should not exist for catalog with product code: {headers}'
					assert (
						'SKU Variante *' not in headers
					), f'SKU Variante * should not exist without a catalog: {headers}'
					assert (
						'Nome do Produto *' in headers
					), f'Nome do Produto * should replace SKU Variante * for catalog without product code: {headers}'
			else:
				assert (
					'Código do Produto *' not in headers
				), f'Código do Produto * should not exist without a catalog: {headers}'
				assert (
					'SKU Variante *' not in headers
				), f'SKU Variante * should not exist without a catalog: {headers}'
				assert (
					'Nome do Produto *' not in headers
				), f'Nome do Produto * should replace SKU Variante * for catalog without product code: {headers}'
			if company.segment.name == service_segment.name:
				assert (
					'Nome do Serviço *' in headers
				), f'Nome do Serviço * should not exist without a catalog: {headers}'

	@pytest.mark.parametrize(
		'indirect_company',
		[
			'company_no_segment_with_catalog_no_codes',
			'company_no_segment_with_catalog_order_code',
			'company_no_segment_with_catalog_product_code',
			'company_no_segment_with_catalog_both_codes',
			'company_with_service_segment_no_catalog_no_codes',
			'company_with_service_segment_no_catalog_order_code',
			'company_with_service_segment_with_catalog_no_codes',
			'company_with_service_segment_with_catalog_order_code',
			'company_with_service_segment_with_catalog_product_code',
			'company_with_service_segment_with_all_codes',
		],
		indirect=True,
	)
	def test_data_validations_exist(self, indirect_company, user):
		"""**Test if data validations are correctly applied in sale order sheet**"""
		company = indirect_company
		generator = ProductAndSaleOrderTemplateGenerator(
			user, company.company_name
		)

		response = generator.get_template()
		wb = load_workbook(BytesIO(response.content))

		sale_order_sheet = wb['Pedidos']
		validations = list(sale_order_sheet.data_validations.dataValidation)

		assert len(validations) > 0

		for validation in validations:
			assert bool(
				validation.formula1
			), f'Validation formula1 is empty: {validation.formula1}'
			assert isinstance(
				validation.formula1, str
			), f'Validation formula1 is not a string: {validation.formula1}'

	@pytest.mark.parametrize(
		'indirect_company',
		[
			'company_no_segment_with_catalog_no_codes',
			'company_no_segment_with_catalog_order_code',
			'company_no_segment_with_catalog_product_code',
			'company_no_segment_with_catalog_both_codes',
			'company_with_service_segment_no_catalog_no_codes',
			'company_with_service_segment_no_catalog_order_code',
			'company_with_service_segment_with_catalog_no_codes',
			'company_with_service_segment_with_catalog_order_code',
			'company_with_service_segment_with_catalog_product_code',
			'company_with_service_segment_with_all_codes',
		],
		indirect=True,
	)
	def test_options_sheet_populated_correctly(
		self,
		indirect_company,
		user,
	):
		"""**Test that the Opções sheet is populated correctly with website names**"""
		company = indirect_company
		generator = ProductAndSaleOrderTemplateGenerator(
			user, company.company_name
		)

		response = generator.get_template()
		wb = load_workbook(BytesIO(response.content))

		options_sheet = wb['Opções']
		options_website_names = [
			cell.value for cell in options_sheet['C'] if cell.value
		]
		obj_website_names = Website.objects.filter(
			website_name__contains=company.company_name
		).values_list('website_humanized_name', flat=True)

		assert len(options_website_names) > 1

		# First options_website_name is the header
		assert len(options_website_names) - 1 == len(obj_website_names)

		for options_website_name in options_website_names[1:]:
			assert options_website_name in obj_website_names

	@pytest.mark.parametrize(
		'indirect_company',
		[
			'company_no_segment_with_catalog_no_codes',
			'company_no_segment_with_catalog_order_code',
			'company_no_segment_with_catalog_product_code',
			'company_no_segment_with_catalog_both_codes',
			'company_with_service_segment_no_catalog_no_codes',
			'company_with_service_segment_no_catalog_order_code',
			'company_with_service_segment_with_catalog_no_codes',
			'company_with_service_segment_with_catalog_order_code',
			'company_with_service_segment_with_catalog_product_code',
			'company_with_service_segment_with_all_codes',
		],
		indirect=True,
	)
	def test_headers_correctness(self, indirect_company, user):
		"""**Test if all headers are correct for 'Produtos' and 'Pedidos' sheets**"""
		company = indirect_company
		generator = ProductAndSaleOrderTemplateGenerator(
			user, company.company_name
		)

		response = generator.get_template()
		wb = load_workbook(BytesIO(response.content))

		# These are the headers that should be always present
		# in the "Produtos" and "Pedidos" sheets
		expected_product_headers = [
			'Preço DE *',
			'Preço POR',
		]
		expected_order_headers = [
			'Data *',
			'CPF/CNPJ',
			'Email',
			'Nome',
			'Telefone do Cliente',
			'Meio de Pagamento',
			'Parcelas',
		]

		if 'Produtos' in wb.sheetnames:
			product_sheet = wb['Produtos']
			headers = [cell.value for cell in product_sheet[2]]
			for expected_header in expected_product_headers:
				assert (
					expected_header in headers
				), f'Product header {expected_header} not found in headers: {headers}'

		if 'Pedidos' in wb.sheetnames:
			order_sheet = wb['Pedidos']
			headers = [cell.value for cell in order_sheet[2]]
			for expected_header in expected_order_headers:
				assert (
					expected_header in headers
				), f'Order header {expected_header} not found in headers: {headers}'

	@pytest.mark.parametrize(
		'indirect_company',
		[
			'company_no_segment_with_catalog_no_codes',
			'company_no_segment_with_catalog_order_code',
			'company_no_segment_with_catalog_product_code',
			'company_no_segment_with_catalog_both_codes',
			'company_with_service_segment_no_catalog_no_codes',
			'company_with_service_segment_no_catalog_order_code',
			'company_with_service_segment_with_catalog_no_codes',
			'company_with_service_segment_with_catalog_order_code',
			'company_with_service_segment_with_catalog_product_code',
			'company_with_service_segment_with_all_codes',
		],
		indirect=True,
	)
	def test_first_row_merged_cells(self, indirect_company, user):
		"""
		**Test that the first row in both 'Produtos' and 'Pedidos' sheets contains
		only merged cells and no two consecutive unmerged cells**
		"""
		company = indirect_company
		generator = ProductAndSaleOrderTemplateGenerator(
			user, company.company_name
		)

		response = generator.get_template()
		wb = load_workbook(BytesIO(response.content))

		for sheet_name in ['Produtos', 'Pedidos']:
			if sheet_name in wb.sheetnames:
				sheet = wb[sheet_name]
				merged_ranges = sheet.merged_cells.ranges

				# Extract all cells in the first row
				first_row_cells = [cell.coordinate for cell in sheet[1]]

				# Ensure no two consecutive cells are unmerged
				unmerged_cells = [
					cell
					for cell in first_row_cells
					if not any(
						cell in merged_range for merged_range in merged_ranges
					)
				]
				for i in range(len(unmerged_cells) - 1):
					assert (
						unmerged_cells[i] not in first_row_cells
						or unmerged_cells[i + 1] not in first_row_cells
					), (
						f'Two consecutive unmerged cells found in the first row of {sheet_name}: '
						f'{unmerged_cells[i]} and {unmerged_cells[i + 1]}'
					)

				assert (
					len(first_row_cells) > 0
				), f'First row in {sheet_name} is empty'
