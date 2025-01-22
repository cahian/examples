from copy import copy
from datetime import datetime
from enum import Enum
from io import BytesIO
from typing import Any, Dict, List, NotRequired, Optional, TypedDict, Union

from django.contrib.auth.models import User
from django.core.exceptions import ObjectDoesNotExist
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet
from rest_framework import serializers
from rest_framework import status as drf_status
from rest_framework.response import Response

from core.api.views.base import APIBatchUpdate
from core.models import (
	Company,
	Seller,
	SpreadsheetIntegrationConfiguration,
	Website,
)
from product.models import ProductSize
from shadow_helpers import (
	normalize_string,
	set_timezone,
	title_case_first,
	validate_positive_numbers,
)

HEADER_GROUP_ROW = 1
HEADER_COLUMN_ROW = 2
TABLE_START_ROW = 3

OPTIONS_HEADER_COLUMN_ROW = 1
OPTIONS_TABLE_START_ROW = 2
OPTIONS_WEBSITE_COLUMN_INDEX = 3

EXCEL_MAX_ROWS = 1048576
EXCEL_ERROR_MESSAGE_MAX_LENGTH = 255

FAKE_DIMENSIONS = ['__dimension_1__', '__dimension_2__', '__dimension_3__']

SpreadsheetCoordinate = Optional[str]
SpreadsheetCoordinateColumn = Optional[int]
SpreadsheetCoordinateRow = int


class SpreadsheetActiveSheet(Enum):
	PRODUCT = 'Produtos'
	SALE_ORDER = 'Pedidos'


class SpreadsheetCoordinatesManager:
	"""
	**Manages spreadsheet column coordinates and provides methods to get and
	update them**

	**Args:**
		* initial_data (Dict[str, SpreadsheetColumn]):
			A dictionary mapping keys to column indices.
	"""

	def __init__(self, initial_data: Dict[str, SpreadsheetCoordinateColumn]):
		self.data = initial_data

	def get_column_index(self, key: str) -> SpreadsheetCoordinateColumn:
		"""
		**Get the column index for a given key**

		**Args:**
			* key (str): The key for which to retrieve the column index.

		**Returns:**
			* Optional[int]: The column index or None if not found.
		"""
		return self.data[key]

	def add_column_index(
		self, key: str, initial_value: SpreadsheetCoordinateColumn
	):
		"""
		**Add a new column with an initial value**

		**Args:**
			* key (str): The key for the new column.
			* initial_value (Optional[int]): The initial column index.

		**Raises:**
			* ValueError: If the column already exists.
		"""
		if key in self.data:
			raise ValueError(f'Column {key} already exists')
		self.data[key] = initial_value

	def update_column_offset(
		self, key: str, value: SpreadsheetCoordinateColumn
	) -> SpreadsheetCoordinateColumn:
		"""
		**Update the column index for a given key**

		**Args:**
			* key (str): The key for the column to update.
			* value (Optional[int]): The value to add to the column index. Use None to reset.

		**Returns:**
			* Optional[int]: The updated column index or None.
		"""
		if self.data[key] is None or value is None:
			self.data[key] = None
		else:
			self.data[key] += value

	def get_coordinate(
		self, key: str, row: SpreadsheetCoordinateRow
	) -> SpreadsheetCoordinate:
		"""
		**Get the spreadsheet coordinate for a given key and row**

		**Args:**
				* key (str): The key for the column.
				* row (int): The row number.

		**Returns:**
				* Optional[str]: The coordinate (e.g., 'A1') or None if the column
				index is missing.
		"""
		col = self.data[key]
		return f'{get_column_letter(col)}{row}' if col else None

	def get_coordinate_range(self, key: str) -> SpreadsheetCoordinate:
		"""
		**Get the coordinate range for a column key**

		**Args:**
			* key (str): The key for the column.

		**Returns:**
			* Optional[str]: The range (e.g., 'A3:A1048576') or None if the column
			index is missing.
		"""
		col = self.data[key]
		return (
			f'{get_column_letter(col)}{TABLE_START_ROW}:{get_column_letter(col)}{EXCEL_MAX_ROWS}'
			if col
			else None
		)


class SpreadsheetColumnNotFound(ValueError):
	pass


class ProductAndSaleOrderBase:
	def __init__(
		self,
		company_name: str,
		active_sheet: SpreadsheetActiveSheet = SpreadsheetActiveSheet.PRODUCT,
	):
		self.company = Company.objects.get(company_name=company_name)
		self.company_name = company_name
		self.is_service_segment = self.company.is_service_segment()
		self.has_catalog = self.company.has_catalog
		self.active_sheet = active_sheet
		try:
			config = SpreadsheetIntegrationConfiguration.objects.get(
				company=self.company
			)
			self.has_product_code = config.has_product_code
			self.has_order_code = config.has_order_code
		except ObjectDoesNotExist:
			self.has_product_code = True
			self.has_order_code = True

	def get_template(self, *args: Any, **kwargs: Any) -> HttpResponse:
		raise NotImplementedError


class ProductAndSaleOrderParamsDict(TypedDict):
	"""
	**A TypedDict representing configuration parameters for processing an XLSX
	spreadsheet in product and sale order contexts**

	**Attributes**:
		* is_service_segment (bool):
			A boolean indicating whether the company is from the service segment.
		* has_catalog (bool):
			A boolean indicating whether the company has a catalog of products/services.
		* has_product_code (bool):
			A boolean indicating whether the spreadsheet contains product codes.
		* has_order_code (bool):
			A boolean indicating whether the spreadsheet contains order codes.
		* format_value (Dict[str, callable]):
			A dictionary that maps column names to formatting functions or data
			types used to format each column in the spreadsheet.
		* required_columns (List[str]):
			A list of columns that are required in the spreadsheet, ensuring that
			essential data fields are always included.
		* validations (List[Dict[str, callable]]):
			A list of dictionaries where each entry defines validation rules for
			specific fields in the spreadsheet, ensuring data consistency and
			correctness.
		* unique_together_fields (Optional[List[Union[List[str], str]]]):
			A list of fields or field combinations that must be unique together,
			preventing duplicate entries based on specific attributes or
			combinations.
		* column_keys (Dict[str, str]):
			A dictionary mapping model field names to their corresponding column
			names in the spreadsheet, allowing for flexible data mapping.
		* ignore_dimensions (Optional[bool]):
			A boolean indicating whether to ignore certain dimensions in the
			spreadsheet. When set to True, dimension columns are omitted.
	"""

	is_service_segment: bool
	has_catalog: bool
	has_product_code: bool
	has_order_code: bool
	format_value: Dict[str, callable]
	required_columns: List[str]
	validations: List[Dict[str, callable]]
	unique_together_fields: NotRequired[List[Union[List[str], str]]]
	column_keys: Dict[str, str]
	ignore_dimensions: NotRequired[bool]


class ProductAndSaleOrderParamsGenerator(ProductAndSaleOrderBase):
	def generate_xlsx_params(
		self, dimensions: Optional[List[str]] = None
	) -> ProductAndSaleOrderParamsDict:
		"""
		**Generates configuration parameters for process an XLSX spreadsheet
		based on company specifications**

		This function defines formatting parameters, validations, required fields,
		and product-specific keys for an XLSX spreadsheet based on the provided
		company name and dimensions list. It customizes the  spreadsheet structure
		based on the company's integration settings and segment (e.g.,'Services'
		or other segments), adjusting fields, validation requirements, and unique
		constraints accordingly.

		**Args:**
			* dimensions (List[str]):
				A list of dimensions (e.g., size, color, or other specific attributes)
				used as columns in the spreadsheet.

		**Returns:**
			* ProductAndSaleOrderParamsDict
		"""
		if dimensions is None:
			dimensions = FAKE_DIMENSIONS
		if self.active_sheet == SpreadsheetActiveSheet.PRODUCT:
			return self._generate_product_xlsx_params(dimensions)
		elif self.active_sheet == SpreadsheetActiveSheet.SALE_ORDER:
			return self._generate_sale_order_xlsx_params()
		else:
			raise ValueError('Invalid active_sheet')

	def _generate_product_xlsx_params(
		self, dimensions: List[str]
	) -> ProductAndSaleOrderParamsDict:
		if not self.has_catalog:
			raise ValueError(
				'Company catalog is required to generate product XLSX params'
			)

		format_value = {
			'SKU *': str,
			'SKU Variante *': str,
			'Nome *': str,
			'Variante *': str,
			'Preço DE *': lambda x: round(float(x), 2),
			'Preço POR': lambda x: round(float(x), 2),
			'Custo Médio Unitário': float,
			dimensions[0]: str,
			dimensions[1]: str,
			dimensions[2]: str,
			'Idade': str,
			'Cor': str,
			'Coleção': str,
			'Estoque Disponível': int,
			'Data do Último Recebimento': lambda x: set_timezone(
				datetime.strptime(x, '%Y-%m-%d %H:%M:%S')
				if isinstance(x, str)
				else x
			),
		}
		required_columns = {
			'SKU *',
			'SKU Variante *',
			'Variante *',
			'Nome *',
			'Preço DE *',
			dimensions[0],
			dimensions[1],
			dimensions[2],
		}
		validations = {
			'Variante *': {
				'validate_func': lambda row: row.get('Variante *'),
				'error_message': lambda row: 'Valor "%s" inválido para "Variante *"'
				% row['Variante *'],
			},
			'Preço DE *': {
				'validate_func': lambda row: validate_positive_numbers(
					row.get('Preço DE *')
				),
				'error_message': lambda row: 'Preço DE "%s" deve ser um número maior que zero'
				% row['Preço DE *'],
			},
			'Preço POR': {
				'validate_func': lambda row: validate_positive_numbers(
					row.get('Preço POR')
				),
				'error_message': lambda row: 'Preço POR "%s" deve ser um número maior que zero'
				% row['Preço POR'],
			},
			'Custo Médio Unitário': {
				'validate_func': lambda row: validate_positive_numbers(
					row.get('Custo Médio Unitário')
				),
				'error_message': lambda row: 'Custo Médio Unitário "%s" deve ser um número maior que zero'
				% row['Custo Médio Unitário'],
			},
		}
		unique_together_fields = [
			['SKU *', 'Variante *'],
			'SKU Variante *',
		]
		column_keys = {
			'supplier_sku_model': 'SKU *',
			'supplier_sku_color': 'SKU *',
			'supplier_sku_size': 'SKU Variante *',
			'size_name': 'Variante *',
			'full_price': 'Preço DE *',
			'special_price': 'Preço POR',
			'gross_cost': 'Custo Médio Unitário',
			'age_name': 'Idade',
			'color_name': 'Cor',
			'season': 'Coleção',
			'product_color_name': 'Nome *',
			'stock': 'Estoque Disponível',
			'last_date': 'Data do Último Recebimento',
		}
		ignore_dimensions = False

		if not self.has_product_code:
			del format_value['SKU *']
			del format_value['SKU Variante *']
			format_value['Nome do Produto *'] = format_value.pop('Nome *')
			del format_value['Variante *']
			del format_value[dimensions[0]]
			del format_value[dimensions[1]]
			del format_value[dimensions[2]]

			required_columns.remove('SKU *')
			required_columns.remove('SKU Variante *')
			required_columns.remove('Nome *')
			required_columns.add('Nome do Produto *')
			required_columns.remove('Variante *')
			required_columns.remove(dimensions[0])
			required_columns.remove(dimensions[1])
			required_columns.remove(dimensions[2])

			del validations['Variante *']

			unique_together_fields = ['Nome do Produto *']

			column_keys['supplier_sku_model'] = 'Nome do Produto *'
			column_keys['supplier_sku_color'] = 'Nome do Produto *'
			column_keys['supplier_sku_size'] = 'Nome do Produto *'
			column_keys['size_name'] = 'Nome do Produto *'
			column_keys['product_color_name'] = 'Nome do Produto *'

			ignore_dimensions = True

		if self.is_service_segment:
			format_value['Custo para o Serviço'] = format_value.pop(
				'Custo Médio Unitário'
			)
			del format_value['Idade']
			del format_value['Cor']
			del format_value['Coleção']
			del format_value['Estoque Disponível']
			del format_value['Data do Último Recebimento']

			validations['Custo para o Serviço'] = {
				'validate_func': lambda row: validate_positive_numbers(
					row.get('Custo para o Serviço')
				),
				'error_message': lambda row: 'Custo para o Serviço "%s" deve ser um número maior que zero'
				% row['Custo para o Serviço'],
			}
			del validations['Custo Médio Unitário']

			column_keys['gross_cost'] = 'Custo para o Serviço'
			del column_keys['age_name']
			del column_keys['color_name']
			del column_keys['season']
			del column_keys['stock']
			del column_keys['last_date']

			if self.has_product_code:  # If the columns was not deleted
				format_value['Código do Produto *'] = format_value.pop('SKU *')
				del format_value['SKU Variante *']
				del format_value['Variante *']

				required_columns.remove('SKU *')
				required_columns.add('Código do Produto *')
				required_columns.remove('SKU Variante *')
				required_columns.remove('Variante *')

				del validations['Variante *']

				unique_together_fields = ['Código do Produto *']

				column_keys['supplier_sku_model'] = 'Código do Produto *'
				column_keys['supplier_sku_color'] = 'Código do Produto *'
				column_keys['supplier_sku_size'] = 'Código do Produto *'
				column_keys['size_name'] = 'Nome *'

		return {
			'has_product_code': self.has_product_code,
			'has_order_code': self.has_order_code,
			'format_value': format_value,
			'required_columns': list(required_columns),
			'validations': list(validations.values()),
			'unique_together_fields': unique_together_fields,
			'column_keys': column_keys,
			'ignore_dimensions': ignore_dimensions,
		}

	def _generate_sale_order_xlsx_params(
		self,
	) -> ProductAndSaleOrderParamsDict:
		format_value = {
			'Numero do Pedido *': str,
			'Canal *': str,
			'Data *': str,
			'Status *': str,
			'SKU Variante *': str,
			'Quantidade *': int,
			'Preço De Unitário': lambda x: round(float(x), 2),
			'Preço Por Unitário': lambda x: round(float(x), 2),
			'Preço Pago Total Produto *': lambda x: round(float(x), 2),
			'CPF/CNPJ': str,
			'Email': str,
			'Nome': str,
			'Telefone do Cliente': str,
			'Frete Total Pedido': lambda x: round(float(x), 2),
			'Meio de Pagamento': str,
			'Parcelas': int,
		}
		required_columns = {
			'Numero do Pedido *',
			'Canal *',
			'Data *',
			'Status *',
			'SKU Variante *',
			'Quantidade *',
			'Preço Pago Total Produto *',
		}
		validations = {
			'Preço De Unitário': {
				'validate_func': lambda row: validate_positive_numbers(
					row.get('Preço De Unitário')
				),
				'error_message': lambda row: 'Preço De Unitário "%s" deve ser um número maior que zero'
				% row['Preço De Unitário'],
			},
			'Preço Por Unitário': {
				'validate_func': lambda row: validate_positive_numbers(
					row.get('Preço Por Unitário')
				),
				'error_message': lambda row: 'Preço Por Unitário "%s" deve ser um número maior que zero'
				% row['Preço Por Unitário'],
			},
			'Preço Pago Total Produto *': {
				'validate_func': lambda row: validate_positive_numbers(
					row.get('Preço Pago Total Produto *')
				),
				'error_message': lambda row: 'Preço Pago Total Produto "%s" deve ser um número maior que zero'
				% row['Preço Pago Total Produto *'],
			},
			'Frete Total Pedido': {
				'validate_func': lambda row: validate_positive_numbers(
					row.get('Frete Total Pedido')
				),
				'error_message': lambda row: 'Frete Total Pedido "%s" deve ser um número maior que zero'
				% row['Frete Total Pedido'],
			},
			'Nome': {
				'validate_func': lambda row: not row.get('Nome')
				or len(row['Nome']) < 255,
				'error_message': lambda row: 'Nome "%s" deve possuir no máximo 255 caracteres'
				% row['Nome'],
			},
			'Email': {
				'validate_func': lambda row: not row.get('Email')
				or len(row['Email']) < 100,
				'error_message': lambda row: 'Email "%s" deve possuir no máximo 100 caracteres'
				% row['Email'],
			},
			'Parcelas': {
				'validate_func': lambda row: validate_positive_numbers(
					row.get('Parcelas')
				),
				'error_message': lambda row: 'Parcelas Pedido "%s" deve ser um número maior que zero'
				% row['Parcelas'],
			},
		}
		column_keys = {
			'sale_order_number': 'Numero do Pedido *',  # aka website_sale_order_id
			'website_name': 'Canal *',
			'date': 'Data *',
			'sale_order_product_size_status': 'Status *',
			'supplier_sku_color': 'SKU Variante *',
			'supplier_sku_size': 'SKU Variante *',
			'quantity': 'Quantidade *',
			'full_price': 'Preço De Unitário',
			'sale_price': 'Preço Por Unitário',
			'total_paid_price': 'Preço Pago Total Produto *',
			'cpf_cnpj': 'CPF/CNPJ',
			'email': 'Email',
			'name': 'Nome',
			'phone': 'Telefone do Cliente',
			'shipping_revenue': 'Frete Total Pedido',
			'order_payment_method_type': 'Meio de Pagamento',
			'order_payment_method_installments': 'Parcelas',
			'is_service_order_item': '__is_service_order_item__',
		}

		websites = Website.objects.filter(
			is_active=True,
			company=self.company,
		).values_list('website_name', flat=True)

		if not self.has_catalog:
			if self.has_product_code:
				raise ValueError(
					"It doesn't make any sense to have a product code without a catalog"
				)
			if not self.is_service_segment:
				raise ValueError(
					"It doesn't make any sense to not have a product catalog if you're not a service company"
				)

			del format_value['SKU Variante *']
			del format_value['Quantidade *']

			required_columns.remove('SKU Variante *')
			required_columns.remove('Quantidade *')

			# a new supplier_sku will be generated in
			# batch_create_or_update_sale_orders.py
			#
			# that's why i'm not deleting it here
			# del column_keys['supplier_sku_color']
			# del column_keys['supplier_sku_size']

			# same logic as comment above
			# del column_keys['quantity']

		if self.has_catalog and not self.has_product_code:
			format_value['Nome do Produto *'] = format_value.pop(
				'SKU Variante *'
			)
			format_value['Quantidade do Produto *'] = format_value.pop(
				'Quantidade *'
			)

			required_columns.remove('SKU Variante *')
			required_columns.add('Nome do Produto *')
			required_columns.remove('Quantidade *')
			required_columns.add('Quantidade do Produto *')

			column_keys['supplier_sku_color'] = 'Nome do Produto *'
			column_keys['supplier_sku_size'] = 'Nome do Produto *'
			column_keys['quantity'] = 'Quantidade do Produto *'

		if not self.has_order_code:
			del format_value['Numero do Pedido *']
			del format_value['Status *']

			required_columns.remove('Numero do Pedido *')
			required_columns.remove('Status *')

			# a new sale_order_number is generated in
			# batch_create_or_update_sale_orders.py::set_sale_order_number
			#
			# that's why i'm not deleting it here
			# del column_keys['sale_order_number']

			# same logic as comment above
			# del column_keys['sale_order_product_size_status']

		if len(websites) == 0:
			raise ValueError('Website names list is empty')
		if len(websites) == 1:
			del format_value['Canal *']

			required_columns.remove('Canal *')

			# a new website_name is generated in
			# batch_create_or_update_sale_orders.py::set_website_name
			#
			# del column_keys['website_name']

		if self.is_service_segment:
			if self.has_catalog:
				if self.has_product_code:
					format_value['Código do Produto *'] = format_value.pop(
						'SKU Variante *'
					)
					format_value['Quantidade do Produto *'] = format_value.pop(
						'Quantidade *'
					)

					required_columns.remove('SKU Variante *')
					required_columns.remove('Quantidade *')
					# the validation of required_columns is more complex in this case,
					# so it's done in
					# batch_create_or_update_sale_orders.py::read_file
					# outside of XLSXDictReader
					#
					# required_columns.add('Código do Produto *')
					# required_columns.add('Quantidade do Produto *')

					column_keys['supplier_sku_color'] = 'Código do Produto *'
					column_keys['supplier_sku_size'] = 'Código do Produto *'
					column_keys['quantity'] = 'Quantidade do Produto *'
				else:
					# the validation of required_columns is more complex in this case,
					# so it's done in
					# batch_create_or_update_sale_orders.py::read_file
					# outside of XLSXDictReader
					required_columns.remove('Nome do Produto *')
					required_columns.remove('Quantidade do Produto *')

			del format_value['Preço De Unitário']
			del format_value['Preço Por Unitário']
			format_value['Preço Pago Total *'] = format_value.pop(
				'Preço Pago Total Produto *'
			)
			del format_value['Frete Total Pedido']
			format_value['Nome do Serviço *'] = str

			required_columns.remove('Preço Pago Total Produto *')
			required_columns.add('Preço Pago Total *')
			# the validation of required_columns is more complex in this case,
			# so it's done in
			# batch_create_or_update_sale_orders.py::read_file
			#
			# required_columns.add('Nome do Serviço *')

			del validations['Preço De Unitário']
			del validations['Preço Por Unitário']
			validations['Preço Pago Total *'] = {
				'validate_func': lambda row: validate_positive_numbers(
					row.get('Preço Pago Total *')
				),
				'error_message': lambda row: 'Preço Pago Total "%s" deve ser um número maior que zero'
				% row['Preço Pago Total *'],
			}
			del validations['Preço Pago Total Produto *']
			del validations['Frete Total Pedido']

			del column_keys['full_price']
			del column_keys['sale_price']
			column_keys['total_paid_price'] = 'Preço Pago Total *'
			del column_keys['shipping_revenue']
			column_keys['product_color_name'] = 'Nome do Serviço *'

		return {
			'is_service_segment': self.is_service_segment,
			'has_catalog': self.has_catalog,
			'has_product_code': self.has_product_code,
			'has_order_code': self.has_order_code,
			'format_value': format_value,
			'required_columns': list(required_columns),
			'validations': list(validations.values()),
			'column_keys': column_keys,
		}


class ProductAndSaleOrderTemplateGenerator(ProductAndSaleOrderBase):
	def __init__(self, user: User, *args: Any, **kwargs: Any) -> None:
		super().__init__(*args, **kwargs)
		self.user = user
		self.website_names = Website.objects.filter(
			is_active=True,
			company=self.company,
		).values_list('website_name', flat=True)

	def get_template(self, *args: Any, **kwargs: Any) -> HttpResponse:
		"""
		**Loads, modifies, and returns an Excel spreadsheet template as an HTTP
		response**

		This function loads a pre-defined Excel file template for product and order
		details, performs necessary updates to the template, and then returns it as
		an HTTP response. It sets the active sheet based on the specified `active_sheet`
		attribute and applies template updates to the product and order sheets.

		**Returns:**
			* HttpResponse
				An HTTP response containing the modified Excel file for download.
		"""
		filename = 'Modelo de Produto e Pedido.xlsx'
		filepath = f'core/static/sheets/{filename}'
		wb = load_workbook(filepath)

		# NOTE: This fixes a bug on Excel tab selection
		wb.active.sheet_view.tabSelected = 0

		wb.active = wb[self.active_sheet.value]

		# Always execute first the update_options_template,
		# because the another update methods depend on the
		# existence of the Opções sheet
		self._update_options_template(wb)

		self._update_product_template(wb)
		self._update_sale_order_template(wb)
		return self._get_spreadsheet_response(wb, filename)

	def _update_options_template(self, wb: Workbook) -> None:
		options_sheet = wb['Opções']

		# 1. Add website name values
		for row, value in enumerate(
			self.website_names, start=OPTIONS_TABLE_START_ROW
		):
			options_sheet.cell(
				row=row, column=OPTIONS_WEBSITE_COLUMN_INDEX, value=value
			)

		# 2. Add supplier sku size values
		# 2.1 From product size model
		supplier_sku_sizes = list(
			ProductSize.objects.filter_user_objects(
				self.user, self.company_name
			)
			.values_list('supplier_sku_size', flat=True)
			.distinct()
		)
		supplier_sku_sizes = [
			sks for sks in supplier_sku_sizes if sks and sks.strip()
		]
		supplier_sku_sizes_length = len(supplier_sku_sizes)

		options_sheet[f'E{OPTIONS_HEADER_COLUMN_ROW}'] = (
			'Produto'  # Set header
		)
		for row_index, value in enumerate(
			supplier_sku_sizes, start=OPTIONS_TABLE_START_ROW
		):
			options_sheet[f'E{row_index}'] = value

		# 2.2 From product template tab
		if supplier_sku_sizes_length > EXCEL_MAX_ROWS:
			raise NotImplementedError(
				f'Companies with more than {EXCEL_MAX_ROWS} products are not supported'
			)

		# 2.2.1 Define SKU list column letter
		sku_list_column_letter = (
			'A'
			if (not self.has_product_code or self.is_service_segment)
			else 'B'
		)

		# 2.2.2 Define options and SKU list start and end row indexes
		options_list_start_row_index = (
			OPTIONS_TABLE_START_ROW + supplier_sku_sizes_length
		)
		options_list_end_row_index = EXCEL_MAX_ROWS

		sku_list_start_row_index = TABLE_START_ROW
		sku_list_end_row_index = TABLE_START_ROW + (
			EXCEL_MAX_ROWS - options_list_start_row_index
		)

		if sku_list_end_row_index > EXCEL_MAX_ROWS:
			diff = EXCEL_MAX_ROWS - sku_list_end_row_index
			options_list_end_row_index += diff
			sku_list_end_row_index += diff

		# 2.2.3 Validate options and SKU list start and end row indexes
		if (
			options_list_end_row_index - options_list_start_row_index
			!= sku_list_end_row_index - sku_list_start_row_index
		):
			raise ValueError(
				f'Row range mismatch between options list and SKU list. '
				f'options list rows span from {options_list_start_row_index} to {options_list_end_row_index} '
				f'({options_list_end_row_index - options_list_start_row_index} rows), while SKU list rows span from '
				f'{sku_list_start_row_index} to {sku_list_end_row_index} '
				f'({sku_list_end_row_index - sku_list_start_row_index} rows). Ensure both lists have the same number of rows.'
			)

		# 2.2.4 Define options and SKU list ranges
		options_list = (
			f'E{options_list_start_row_index}:E{options_list_end_row_index}'
		)
		sku_list = f'Produtos!{sku_list_column_letter}{sku_list_start_row_index}:{sku_list_column_letter}{sku_list_end_row_index}'

		# 2.2.5 Define ArrayFormula
		options_sheet[f'E{options_list_start_row_index}'] = ArrayFormula(
			ref=options_list,
			text=f'=IF({sku_list}="", "", {sku_list})',
		)

		# 3. Hide the entire sheet
		options_sheet.sheet_state = 'hidden'

	def _update_product_template(self, wb: Workbook) -> None:
		product_sheet = wb['Produtos']

		# 1. Remove this worksheet
		if not self.has_catalog:  # If company has no catalog
			wb.remove(product_sheet)
			return

		# 2. Dynamic Columns
		if not self.has_product_code:
			self._delete_column_with_merged_ranges(product_sheet, 'SKU *')
			self._delete_column_with_merged_ranges(
				product_sheet, 'SKU Variante *'
			)
			self._rename_column_header(
				product_sheet, 'Nome *', 'Nome do Produto *'
			)
			self._delete_column_with_merged_ranges(product_sheet, 'Variante *')
			self._delete_column_with_merged_ranges(
				product_sheet, 'Departamento *'
			)
			self._delete_column_with_merged_ranges(
				product_sheet, 'Categoria *'
			)
			self._delete_column_with_merged_ranges(
				product_sheet, 'Subcategoria *'
			)

		if self.is_service_segment:
			self._delete_column_with_merged_ranges(product_sheet, 'Idade')
			self._delete_column_with_merged_ranges(product_sheet, 'Cor')
			self._delete_column_with_merged_ranges(product_sheet, 'Coleção')
			self._delete_column_with_merged_ranges(
				product_sheet, 'Estoque Disponível'
			)
			self._delete_column_with_merged_ranges(
				product_sheet, 'Data do Último Recebimento'
			)
			if self.has_product_code:  # If the columns was not deleted
				self._rename_column_header(
					product_sheet, 'SKU *', 'Código do Produto *'
				)
				self._delete_column_with_merged_ranges(
					product_sheet, 'SKU Variante *'
				)
				self._delete_column_with_merged_ranges(
					product_sheet, 'Variante *'
				)
			self._rename_column_header(
				product_sheet, 'Custo Médio Unitário', 'Custo para o Serviço'
			)

		# 3. Update dimensions columns names
		if self.has_product_code:  # If the columns was not deleted
			seller = Seller.get_seller_from_company_name(self.company_name)
			seller_humanized_dimensions = [
				f'{seller.get_humanized_dimension_name(dimension)} *'
				for dimension in range(1, 4)
			]
			self._rename_column_header(
				product_sheet, 'Departamento *', seller_humanized_dimensions[0]
			)
			self._rename_column_header(
				product_sheet, 'Categoria *', seller_humanized_dimensions[1]
			)
			self._rename_column_header(
				product_sheet, 'Subcategoria *', seller_humanized_dimensions[2]
			)

		# 4. Update columns format
		try:
			date_column_letter = get_column_letter(
				self._find_header_index_by_value(
					product_sheet, 'Data do Último Recebimento'
				)
			)
		except SpreadsheetColumnNotFound:
			date_column_letter = None
		for column_letter, column in product_sheet.column_dimensions.items():
			if date_column_letter and column_letter == date_column_letter:
				column.number_format = 'd/m/yyyy'
			else:
				column.number_format = 'General'

		# 5. Adjust column width
		self._adjust_column_width(product_sheet)

	def _update_sale_order_template(self, wb: Workbook) -> None:
		sale_order_sheet = wb['Pedidos']

		# 1. Dynamic Columns
		coord = SpreadsheetCoordinatesManager(
			{
				'website': 2,
				'status': 4,
				'sku': 5,
			}
		)

		if not self.has_catalog:
			if self.has_product_code:
				raise ValueError(
					"It doesn't make any sense to have a product code without a catalog"
				)
			if not self.is_service_segment:
				raise ValueError(
					"It doesn't make any sense to not have a catalog if you're not a service company"
				)

			self._delete_column_with_merged_ranges(
				sale_order_sheet, 'SKU Variante *'
			)
			self._delete_column_with_merged_ranges(
				sale_order_sheet, 'Quantidade *'
			)

			coord.update_column_offset('sku', None)

		if self.has_catalog and not self.has_product_code:
			self._rename_column_header(
				sale_order_sheet, 'SKU Variante *', 'Nome do Produto *'
			)
			self._rename_column_header(
				sale_order_sheet, 'Quantidade *', 'Quantidade do Produto *'
			)

		if not self.has_order_code:
			self._delete_column_with_merged_ranges(
				sale_order_sheet, 'Numero do Pedido *'
			)
			self._delete_column_with_merged_ranges(
				sale_order_sheet, 'Status *'
			)

			# Update data validations columns coordinates
			coord.update_column_offset('website', -1)
			coord.update_column_offset('status', None)
			coord.update_column_offset('sku', -2)

		if len(self.website_names) == 0:
			raise ValueError('Website names list is empty')
		if len(self.website_names) == 1:
			self._delete_column_with_merged_ranges(sale_order_sheet, 'Canal *')

			# Update data validations columns coordinates
			coord.update_column_offset('website', None)
			coord.update_column_offset('status', -1)
			coord.update_column_offset('sku', -1)

		if self.is_service_segment:
			if self.has_catalog:
				if self.has_product_code:
					self._rename_column_header(
						sale_order_sheet,
						'SKU Variante *',
						'Código do Produto *',
					)
					self._rename_column_header(
						sale_order_sheet,
						'Quantidade *',
						'Quantidade do Produto *',
					)
				coord.add_column_index(
					'quantity',
					self._find_header_index_by_value(
						sale_order_sheet, 'Quantidade do Produto *'
					),
				)

			self._delete_column_with_merged_ranges(
				sale_order_sheet, 'Preço De Unitário'
			)
			self._delete_column_with_merged_ranges(
				sale_order_sheet, 'Preço Por Unitário'
			)
			self._rename_column_header(
				sale_order_sheet,
				'Preço Pago Total Produto *',
				'Preço Pago Total *',
			)
			self._delete_column_with_merged_ranges(
				sale_order_sheet, 'Frete Total Pedido'
			)
			self._add_column_with_merged_ranges(
				sale_order_sheet,
				'Preço Pago Total *',
				'Nome do Serviço *',
			)
			coord.add_column_index(
				'service',
				self._find_header_index_by_value(
					sale_order_sheet, 'Nome do Serviço *'
				),
			)

			# No need to update data validations columns coordinates, as the deleted
			# columns, 'Quantidade *', 'Preço De Unitário' and 'Frete Total Pedido',
			# are located after the data validations coordinates.

		# 2. Data Validations
		# 2.1 Website Validation
		if coord.get_column_index('website'):
			website_error = 'O canal informado não foi encontrado.'
			website_options = f' Escolha um dos canais disponíveis: {', '.join(self.website_names)}.'
			if (
				len(website_error + website_options)
				<= EXCEL_ERROR_MESSAGE_MAX_LENGTH
			):
				website_error += website_options
			website_validation = DataValidation(
				type='list',
				formula1=f'Opções!$C${OPTIONS_TABLE_START_ROW}:$C${OPTIONS_HEADER_COLUMN_ROW + len(self.website_names)}',
				showErrorMessage=True,
				allowBlank=True,
				errorStyle='stop',
				errorTitle='Canal não encontrado',
				error=website_error,
			)
			sale_order_sheet.add_data_validation(website_validation)
			website_validation.add(coord.get_coordinate_range('website'))

		# 2.2 Status Validation
		if coord.get_column_index('status'):
			options_sheet = wb['Opções']
			status_range = f'$A${OPTIONS_TABLE_START_ROW}:$A${OPTIONS_HEADER_COLUMN_ROW + 3}'
			status_options = self._extract_values_from_range(
				options_sheet, status_range
			)
			status_validation = DataValidation(
				type='list',
				formula1=f'Opções!{status_range}',
				showErrorMessage=True,
				allowBlank=True,
				errorStyle='stop',
				errorTitle='Status inválido',
				error=(
					'O status informado não é válido. '
					f'Escolha um dos seguintes opções: {', '.join(status_options)}.'
				),
			)
			sale_order_sheet.add_data_validation(status_validation)
			status_validation.add(coord.get_coordinate_range('status'))

		# 2.3 SKU Validation
		if coord.get_column_index('sku'):
			sku_list = (
				f'Opções!$E${OPTIONS_TABLE_START_ROW}:$E${EXCEL_MAX_ROWS}'
			)

			if self.is_service_segment and self.has_catalog:
				blank_list = f'Opções!$A${EXCEL_MAX_ROWS}'

				sku_header_name = self._find_header_name_by_column_key(
					sale_order_sheet,
					SpreadsheetActiveSheet.SALE_ORDER,
					'supplier_sku_size',
				)
				quantity_header_name = self._find_header_name_by_column_key(
					sale_order_sheet,
					SpreadsheetActiveSheet.SALE_ORDER,
					'quantity',
				)
				service_header_name = self._find_header_name_by_column_key(
					sale_order_sheet,
					SpreadsheetActiveSheet.SALE_ORDER,
					'product_color_name',
				)

				sku_validation = DataValidation(
					type='list',
					formula1=(
						f'IF(ISBLANK(INDIRECT(ADDRESS(ROW(),COLUMN()+2,4))), {sku_list}, {blank_list})'
					),
					showErrorMessage=True,
					allowBlank=True,
					errorStyle='stop',
					errorTitle=f'{title_case_first(sku_header_name)} não permitido',
					# The maximum allowed length for the error message is 255 characters.
					error=(
						f'O {sku_header_name.lower()} não é permitido. Verifique se:\n'
						f'1. O {sku_header_name.lower()} informado está na lista de produtos.\n'
						f'2. Cada linha contém apenas um dos dois: o {service_header_name.lower()}, ou o {sku_header_name.lower()} e a {quantity_header_name.lower()} correspondente.'
					),
				)
				sale_order_sheet.add_data_validation(sku_validation)
				sku_validation.add(coord.get_coordinate_range('sku'))

				quantity_validation = DataValidation(
					type='custom',
					formula1=(
						'=AND(ISNUMBER(INDIRECT(ADDRESS(ROW(),COLUMN(),4))), ISBLANK(INDIRECT(ADDRESS(ROW(),COLUMN()+1,4))))'
					),
					showErrorMessage=True,
					allowBlank=True,
					errorStyle='stop',
					errorTitle='Produto e serviço não podem ser preenchidos no mesmo pedido',
					error=(
						f'Cada linha deve conter apenas um dos dois: o {sku_header_name.lower()} e a {quantity_header_name.lower()}, '
						f'ou o {service_header_name.lower()}. Não é permitido preencher ambos na mesma linha.'
					),
				)
				sale_order_sheet.add_data_validation(quantity_validation)
				quantity_validation.add(coord.get_coordinate_range('quantity'))

				service_validation = DataValidation(
					type='custom',
					formula1=(
						'=AND(ISBLANK(INDIRECT(ADDRESS(ROW(),COLUMN()-1,4))), ISBLANK(INDIRECT(ADDRESS(ROW(),COLUMN()-2,4))))'
					),
					showErrorMessage=True,
					allowBlank=True,
					errorStyle='stop',
					errorTitle='Produto e serviço não podem ser preenchidos no mesmo pedido',
					error=(
						f'Cada linha deve conter apenas um dos dois: o {sku_header_name.lower()} e a {quantity_header_name.lower()}, '
						f'ou o {service_header_name.lower()}. Não é permitido preencher ambos na mesma linha.'
					),
				)
				sale_order_sheet.add_data_validation(service_validation)
				service_validation.add(coord.get_coordinate_range('service'))
			else:
				sku_validation = DataValidation(
					type='list',
					formula1=sku_list,
					showErrorMessage=True,
					allowBlank=True,
					errorStyle='stop',
					errorTitle='Produto não encontrado',
					error='O produto informado não foi encontrado.',
				)
				sale_order_sheet.add_data_validation(sku_validation)
				sku_validation.add(coord.get_coordinate_range('sku'))

		# 3. Update default values
		if coord.get_column_index('sku'):
			product_sheet = wb['Produtos']
			column_letter = self._find_header_letter_by_column_key(
				product_sheet,
				SpreadsheetActiveSheet.PRODUCT,
				'supplier_sku_size',
			)
			product_coord = f'{column_letter}{TABLE_START_ROW}'
			product_cell = product_sheet[product_coord]
			sale_order_sheet[coord.get_coordinate('sku', TABLE_START_ROW)] = (
				product_cell.value
			)

		# 4. Update columns format
		date_column_letter = get_column_letter(
			self._find_header_index_by_value(sale_order_sheet, 'Data *')
		)
		for (
			column_letter,
			column,
		) in sale_order_sheet.column_dimensions.items():
			if column_letter == date_column_letter:
				column.number_format = 'd/m/yyyy'
			else:
				column.number_format = 'General'

		# 5. Adjust column width
		self._adjust_column_width(sale_order_sheet)

	@classmethod
	def _add_column_with_merged_ranges(
		cls, ws: Worksheet, column_index_or_name: Union[int, str], value: str
	) -> None:
		"""
		**Add a column and handle merged cells**

		This method adds a new column to the specified index or name in the
		worksheet. It shifts existing columns to the right to make space for the
		new column, and it also handles any merged cells that may be affected by
		this change.

		**Args:**
			* ws (Worksheet):
				A worksheet object.
			* column_index_or_name (int | str):
				The index or name of the column to be added (1-based).
			* value (str):
				The value to be set in the new column header.
		"""
		if isinstance(column_index_or_name, int):
			index = column_index_or_name
		elif isinstance(column_index_or_name, str):
			index = cls._find_header_index_by_value(ws, column_index_or_name)
		else:
			raise ValueError('column_index_or_name must be an int or str')

		# 1. Collect existing single-row merged cells
		old_ranges = []
		merged_cells_coords = set()
		for merged_range in ws.merged_cells.ranges:
			min_col, min_row, max_col, max_row = range_boundaries(
				str(merged_range)
			)
			if min_row != max_row:
				raise ValueError('Only single row merged cells are supported')
			old_ranges.append(
				(merged_range, min_col, max_col, min_row, max_row)
			)

			# Track every cell coordinate in this merged range
			for row_idx in range(min_row, max_row + 1):
				for col_idx in range(min_col, max_col + 1):
					merged_cells_coords.add((row_idx, col_idx))

		# 2. Include every non-merged cell in the first row as a "merged range"
		first_row = 1
		max_col_in_sheet = ws.max_column
		for col_idx in range(1, max_col_in_sheet + 1):
			if (first_row, col_idx) not in merged_cells_coords:
				# Treat this single cell as a merged range
				single_cell_range_str = f'{get_column_letter(col_idx)}{first_row}:{get_column_letter(col_idx)}{first_row}'
				old_ranges.append(
					(
						single_cell_range_str,
						col_idx,
						col_idx,
						first_row,
						first_row,
					)
				)

		# 3. Handle merged ranges
		new_ranges = []
		for merged_range, min_col, max_col, min_row, max_row in old_ranges:
			if max_col >= index:
				# The top-left cell data is crucial in a merged cell range because
				# it holds the primary value and style information for the entire
				# merged range. When cells are merged, the content and formatting
				# from the top-left cell are displayed across the merged area,
				# while the other cells in the range are effectively blank and just
				# part of the merged visual representation.
				top_left_cell_data = cls._get_cell_data(ws, min_row, min_col)

				if min_col == max_col:
					if min_col == index:
						max_col += 1
					else:
						min_col += 1
						max_col += 1
				else:
					ws.unmerge_cells(str(merged_range))

					if min_col > index:
						min_col += 1
					elif min_col == index:
						if min_row != max_row:
							raise ValueError(
								'Only single row merged cells are supported'
							)
						cls._move_cell_right(ws, min_col, min_row)
					if max_col >= index:
						max_col += 1

				new_ranges.append(
					{
						'top_left_cell_data': top_left_cell_data,
						'min_col': min_col,
						'max_col': max_col,
						'min_row': min_row,
						'max_row': max_row,
					}
				)

		# 4. Shift existing columns to the right
		ws.insert_cols(index)

		# 5. Set the header for the new column
		cell = ws.cell(row=HEADER_COLUMN_ROW, column=index)
		cell.value = value
		cell.alignment = Alignment(horizontal='center')
		cell.border = Border(
			left=Side(style='thin'),
			right=Side(style='thin'),
			top=Side(style='thin'),
			bottom=Side(style='thin'),
		)
		cell.font = Font(
			bold=True, color='E97132' if value.endswith('*') else None
		)

		# 6. Re-apply merges for shifted ranges
		for range_data in new_ranges:
			top_left_cell_data, min_col, max_col, min_row, max_row = (
				range_data['top_left_cell_data'],
				range_data['min_col'],
				range_data['max_col'],
				range_data['min_row'],
				range_data['max_row'],
			)

			if min_row != max_row:
				raise ValueError('Only single row merged cells are supported')

			cell = ws.cell(row=min_row, column=min_col)
			cls._copy_cell(cell, top_left_cell_data)
			if min_col != max_col:
				new_range = f'{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}'
				ws.merge_cells(new_range)

	@classmethod
	def _delete_column_with_merged_ranges(
		cls, ws: Worksheet, column_index_or_name: Union[int, str]
	) -> None:
		"""
		**Delete a column and handle merged cells**

		**Args:**
			* ws (Worksheet):
				A worksheet object.
			* column_index_or_name (int | str):
				The index or name of the column to be deleted (1-based).
		"""
		if isinstance(column_index_or_name, int):
			index = column_index_or_name
		elif isinstance(column_index_or_name, str):
			index = cls._find_header_index_by_value(ws, column_index_or_name)
		else:
			raise ValueError('column_index_or_name must be an int or str')

		old_ranges = []
		for merged_range in ws.merged_cells.ranges:
			min_col, min_row, max_col, max_row = range_boundaries(
				str(merged_range)
			)
			if min_row != max_row:
				raise ValueError('Only single row merged cells are supported')
			old_ranges.append(
				(merged_range, min_col, max_col, min_row, max_row)
			)

		new_ranges = []
		for merged_range, min_col, max_col, min_row, max_row in old_ranges:
			if max_col >= index:
				# The top-left cell data is crucial in a merged cell range because
				# it holds the primary value and style information for the entire
				# merged range. When cells are merged, the content and formatting
				# from the top-left cell are displayed across the merged area,
				# while the other cells in the range are effectively blank and just
				# part of the merged visual representation.
				top_left_cell_data = cls._get_cell_data(ws, min_row, min_col)

				ws.unmerge_cells(str(merged_range))

				if min_col == max_col:
					if min_col != index:
						min_col -= 1
						max_col -= 1
				else:
					if min_col > index:
						min_col -= 1
					elif min_col == index:
						if min_row != max_row:
							raise ValueError(
								'Only single row merged cells are supported'
							)
						cls._move_cell_right(ws, min_col, min_row)
					if max_col >= index:
						max_col -= 1

				new_ranges.append(
					{
						'top_left_cell_data': top_left_cell_data,
						'min_col': min_col,
						'max_col': max_col,
						'min_row': min_row,
						'max_row': max_row,
					}
				)

		ws.delete_cols(index)

		for range_data in new_ranges:
			top_left_cell_data, min_col, max_col, min_row, max_row = (
				range_data['top_left_cell_data'],
				range_data['min_col'],
				range_data['max_col'],
				range_data['min_row'],
				range_data['max_row'],
			)

			if min_row != max_row:
				raise ValueError('Only single row merged cells are supported')

			cell = ws.cell(row=min_row, column=min_col)
			cls._copy_cell(cell, top_left_cell_data)
			if min_col != max_col:
				new_range = f'{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}'
				ws.merge_cells(new_range)

	@classmethod
	def _rename_column_header(
		cls,
		ws: Worksheet,
		column_index_or_name: Union[int, str],
		value: str,
	) -> None:
		"""
		**Rename a column header**

		**Args:**
			* ws (Worksheet):
				The worksheet object from openpyxl.
			* column_index_or_name (int or str):
				The index or name of the column to be renamed (1-based).
			* value (str):
				The new name of the column.
		"""
		if isinstance(column_index_or_name, int):
			index = column_index_or_name
		elif isinstance(column_index_or_name, str):
			index = cls._find_header_index_by_value(ws, column_index_or_name)
		else:
			raise ValueError('column_index_or_name must be an int or str')

		ws.cell(row=HEADER_COLUMN_ROW, column=index).value = value

	@staticmethod
	def _extract_values_from_range(ws: Worksheet, range_str: str) -> List[str]:
		"""
		**Extract values from a given range in an Excel worksheet**

		**Args:**
			* ws (Worksheet):
				The Excel worksheet where the range is located.
			* range_str (str):
				The range in Excel format (e.g., "A$2:$A$4").

		**Returns:**
			* List[str]:
				A list of non-empty values extracted from the specified range.
		"""
		min_col, min_row, max_col, max_row = range_boundaries(range_str)
		return [
			ws.cell(row=row, column=min_col).value
			for row in range(min_row, max_row + 1)
			for col in range(min_col, max_col + 1)
			if ws.cell(row=row, column=col).value is not None
		]

	@staticmethod
	def _adjust_column_width(ws: Worksheet) -> None:
		"""
		**Adjust the column width of the worksheet**

		**Args:**
			* ws (Worksheet):
				The worksheet object from openpyxl.
		"""
		for col in ws.columns:
			max_length = 0
			col_letter = get_column_letter(col[0].column)
			for cell in col:
				try:
					if len(str(cell.value)) > max_length:
						max_length = len(str(cell.value))
				except:
					pass
			adjusted_width = max_length + 2
			ws.column_dimensions[col_letter].width = adjusted_width

	def _find_header_name_by_column_key(
		self,
		ws: Worksheet,
		active_sheet: SpreadsheetActiveSheet,
		column_key: str,
	) -> str:
		"""
		**Find the header name for a given column key using the column letter**

		**Args:**
			ws (Worksheet):
				The worksheet object from openpyxl.
			active_sheet (SpreadsheetActiveSheet):
				The active sheet object containing metadata about the spreadsheet.
			column_key (str):
				The key representing the column to locate.

		**Returns:**
			* str:
				The header value of the column corresponding to the given key.

		**Raises:**
			* SpreadsheetColumnNotFound:
				If the column key is not found in the sheet.
		"""
		column_letter = self._find_header_letter_by_column_key(
			ws, active_sheet, column_key
		)
		header_cell = ws[f'{column_letter}{HEADER_COLUMN_ROW}']
		header_value = header_cell.value
		header_name = (
			header_value[:-2] if header_value.endswith(' *') else header_value
		)
		return header_name

	def _find_header_letter_by_column_key(
		self,
		ws: Worksheet,
		active_sheet: SpreadsheetActiveSheet,
		column_key: str,
	) -> int:
		"""
		**Find the header letter for a given column key**

		**Args:**
			* ws (Worksheet):
				The worksheet object from openpyxl.
			* active_sheet (SpreadsheetActiveSheet):
				The active sheet object containing metadata about the spreadsheet.
			* column_key (str):
				The key representing the column to locate.

		**Returns:**
			* str:
				The letter of the column corresponding to the given key.

		**Raises:**
			* SpreadsheetColumnNotFound:
				If the column key is not found in the sheet.
		"""
		return get_column_letter(
			self._find_header_index_by_column_key(ws, active_sheet, column_key)
		)

	def _find_header_index_by_column_key(
		self,
		ws: Worksheet,
		active_sheet: SpreadsheetActiveSheet,
		column_key: str,
	) -> int:
		"""
		**Find the header index for a given column key**

		**Args:**
			* ws (Worksheet):
				The worksheet object from openpyxl.
			* active_sheet (SpreadsheetActiveSheet):
				The active sheet object containing metadata about the spreadsheet.
			* column_key (str):
				The key representing the column to locate.

		**Returns:**
			* int:
				The index (1-based) of the column corresponding to the given key.

		**Raises:**
			* SpreadsheetColumnNotFound:
				If the column key is not found in the sheet.
		"""
		params = ProductAndSaleOrderParamsGenerator(
			self.company_name, active_sheet
		).generate_xlsx_params()
		column_keys = params['column_keys']
		return self._find_header_index_by_value(ws, column_keys[column_key])

	@classmethod
	def _find_header_index_by_value(cls, ws: Worksheet, value: str) -> int:
		"""
		**Find the header index by cell value**

		**Args:**
			* ws (Worksheet):
				The worksheet object from openpyxl.
			* value (str):
				The value to search for.

		**Returns:**
			* int:
				The column index (1-based) of the cell with the value.

		**Raises:**
			* SpreadsheetColumnNotFound:
				If the column with the value is not found.
		"""
		return cls._find_column_index_by_value(ws, HEADER_COLUMN_ROW, value)

	@staticmethod
	def _find_column_index_by_value(
		ws: Worksheet, row: int, value: str
	) -> int:
		"""
		**Find the column index by cell value**

		**Args:**
			* ws (Worksheet):
				The worksheet object from openpyxl.
			* row (int):
				The row index (1-based) of the cell with the value.
			* value (str):
				The value to search for. Example: 'Status *'.

		**Returns:**
			* int:
				The column index (1-based) of the cell with the value.

		**Raises:**
			* SpreadsheetColumnNotFound:
				If the column with the value is not found.
		"""
		for row in ws.iter_rows(
			min_row=HEADER_COLUMN_ROW, max_row=HEADER_COLUMN_ROW, min_col=1
		):
			for cell in row:
				if normalize_string(cell.value) == normalize_string(value):
					return cell.column
		raise SpreadsheetColumnNotFound(
			f'Column with value "{value}" not found'
		)

	@staticmethod
	def _get_cell_data(ws: Worksheet, row: int, column: int) -> Dict[str, Any]:
		"""
		**Get cell data**

		**Args:**
			* ws (Worksheet):
				The worksheet object from openpyxl.
			* row (int):
				The row index (1-based) of the cell.
			* column (int):
				The column index (1-based) of the cell.

		**Returns:**
			* dict:
				A dictionary containing the cell data.
		"""
		cell = ws.cell(row=row, column=column)
		return {
			'value': copy(cell.value),
			'alignment': copy(cell.alignment),
			'fill': copy(cell.fill),
			'border': copy(cell.border),
			'font': copy(cell.font),
		}

	@staticmethod
	def _copy_cell(
		target_cell: Cell, source_cell_data: Dict[str, Any]
	) -> None:
		"""
		**Copy cell data to another cell**

		**Args:**
			* target_cell (Cell):
				The target cell object from openpyxl.
			* source_cell_data (dict):
				The source cell data.
		"""
		target_cell.value = source_cell_data['value']
		target_cell.alignment = source_cell_data['alignment']
		target_cell.fill = source_cell_data['fill']
		target_cell.border = source_cell_data['border']
		target_cell.font = source_cell_data['font']

	@staticmethod
	def _move_cell_right(ws: Worksheet, col: int, row: int) -> None:
		"""
		**Move the content from one cell to the next cell to the right**

		**Args:**
			* ws (Worksheet):
				The worksheet object from openpyxl.
			* col (int):
				The column index of the cell to move (1-based).
			* row (int):
				The row index of the cell to move (1-based).
		"""
		coordinate1 = f'{get_column_letter(col)}{row}'
		coordinate2 = f'{get_column_letter(col + 1)}{row}'
		ws[coordinate2].value = ws[coordinate1].value
		ws[coordinate1].value = None

	@staticmethod
	def _get_spreadsheet_response(wb: Workbook, filename: str) -> HttpResponse:
		with BytesIO() as buffer:
			wb.save(buffer)
			buffer.seek(0)

			response = HttpResponse(
				buffer.getvalue(),
				content_type='application/ms-excel',
			)
			response['Content-Disposition'] = (
				f'attachment; filename="{filename}"'
			)
			response['Access-Control-Expose-Headers'] = 'Content-Disposition'

		return response


class WriteProductAndSaleOrderSerializer(serializers.Serializer):
	upload = serializers.FileField()


class BatchWriteProductAndSaleOrder(APIBatchUpdate):
	permission_required = 'user.can_batch_write_product_and_sale_order'
	serializer_class = WriteProductAndSaleOrderSerializer
	template_path = 'core/static/'
	template_name = 'Modelo de Produto e Pedido.xlsx'

	def get_serializer(
		self, *args, **kwargs
	) -> WriteProductAndSaleOrderSerializer:
		return self.serializer_class()

	def post(self, *args, **kwargs) -> Response:
		company = Company.objects.filter_user_objects(
			self.request.user, self.kwargs['company_name']
		).first()
		seller_id = Seller.get_seller_from_company_name(
			self.kwargs['company_name'],
		).pk

		input_params = {
			'seller_id': seller_id,
			'company_name': self.kwargs['company_name'],
			'company_humanized_name': company.company_humanized_name,
			'company_id': company.pk,
		}

		# File
		input_file = self.request.data['upload']

		# AsyncProcess
		dict_return = self.process_async(
			process_type='batch_write_products_and_sale_orders',
			input_params=input_params,
			input_file=input_file,
			sheet_names=(
				[
					SpreadsheetActiveSheet.PRODUCT.value,
					SpreadsheetActiveSheet.SALE_ORDER.value,
				]
				if company.has_catalog
				else [SpreadsheetActiveSheet.SALE_ORDER.value]
			),
		)

		if dict_return['status'] == 'error':
			return Response(
				dict_return, status=drf_status.HTTP_400_BAD_REQUEST
			)
		else:
			return Response(dict_return)

	def get_template(self, *args, **kwargs) -> HttpResponse:
		return ProductAndSaleOrderTemplateGenerator(
			self.request.user, self.kwargs['company_name']
		).get_template(*args, **kwargs)
